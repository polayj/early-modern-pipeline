#!/usr/bin/env python3
"""
01_parse_import_records.py
Parse all 57 xlsx import records (1696–1755) into normalized CSV, JSON,
and per-location markdown documents for LightRAG ingestion.

Usage:
    python pipeline/01_parse_import_records.py
    python pipeline/01_parse_import_records.py --input import_records/ --output output/import_records_parsed/

Outputs:
    output/import_records_parsed/all_records.csv
    output/import_records_parsed/all_records.json
    output/import_records_parsed/import_docs/<YYYY-YYYY>_<LOCATION>.md
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from tqdm import tqdm

# ── Column index mapping (0-based) ────────────────────────────────────────────
# Header: YEAR | COMMODITY_NAME | FROM_LOCATION | TO_LOCATION | ton | hh |
#         quarter | lb | (None) | Total Pounds | (None) | WEIGHT | ...
COL_YEAR = 0
COL_COMMODITY = 1
COL_FROM = 2
COL_TO = 3
COL_TON = 4
COL_HH = 5
COL_QUARTER = 6
COL_LB = 7
COL_TOTAL_POUNDS = 9
COL_WEIGHT = 11

SKIP_SHEETS = {"Master Goods Sheet"}


def extract_years_from_filename(fname: str) -> tuple[str, str]:
    """Extract start/end years from '1696-1697 British Commodity Imports.xlsx'."""
    m = re.match(r"(\d{4})-(\d{4})", fname)
    if m:
        return m.group(1), m.group(2)
    # Fallback: look for any 4-digit year
    years = re.findall(r"\d{4}", fname)
    if len(years) >= 2:
        return years[0], years[1]
    if len(years) == 1:
        return years[0], years[0]
    return "unknown", "unknown"


def safe_float(val) -> float | None:
    """Convert a cell value to float, returning None for blank/zero."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f != 0.0 else None
    except (TypeError, ValueError):
        return None


def parse_sheet(ws, year_start: str, year_end: str, filename: str) -> list[dict]:
    """Parse one location sheet into a list of record dicts."""
    records = []
    header_row = None
    year_str = f"{year_start}-{year_end}"

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # First row is always the header
        if row_idx == 0:
            header_row = [str(c).strip() if c is not None else "" for c in row]
            continue

        # Skip fully empty rows
        if all(c is None for c in row):
            continue

        # Commodity name (col 1) must be present
        commodity = row[COL_COMMODITY] if len(row) > COL_COMMODITY else None
        if commodity is None:
            continue
        commodity = str(commodity).strip()
        if not commodity:
            continue

        # Year: prefer cell value, fall back to filename year
        year_val = row[COL_YEAR] if len(row) > COL_YEAR else None
        if year_val is not None:
            try:
                year_cell = str(int(float(year_val)))
            except (TypeError, ValueError):
                year_cell = year_start
        else:
            year_cell = year_start

        from_loc = str(row[COL_FROM]).strip() if len(row) > COL_FROM and row[COL_FROM] else ws.title
        to_loc = str(row[COL_TO]).strip() if len(row) > COL_TO and row[COL_TO] else "London"

        records.append({
            "file": filename,
            "year_range": year_str,
            "year": year_cell,
            "year_start": year_start,
            "year_end": year_end,
            "commodity_name": commodity,
            "from_location": from_loc,
            "to_location": to_loc,
            "ton": safe_float(row[COL_TON] if len(row) > COL_TON else None),
            "hh": safe_float(row[COL_HH] if len(row) > COL_HH else None),
            "quarter": safe_float(row[COL_QUARTER] if len(row) > COL_QUARTER else None),
            "lb": safe_float(row[COL_LB] if len(row) > COL_LB else None),
            "total_pounds": safe_float(row[COL_TOTAL_POUNDS] if len(row) > COL_TOTAL_POUNDS else None),
            "weight_tons": safe_float(row[COL_WEIGHT] if len(row) > COL_WEIGHT else None),
        })

    return records


def parse_workbook(xlsx_path: Path) -> list[dict]:
    """Parse all data sheets in one xlsx workbook."""
    fname = xlsx_path.name
    year_start, year_end = extract_years_from_filename(fname)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        records = parse_sheet(ws, year_start, year_end, fname)
        all_records.extend(records)

    wb.close()
    return all_records


def build_markdown_docs(records: list[dict], docs_dir: Path) -> int:
    """
    Write one markdown doc per (year_range, from_location) pair.
    Returns number of docs written.
    """
    # Group by (year_range, from_location)
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for r in records:
        # Only include rows that actually have data (non-zero weight or pounds)
        if r["total_pounds"] or r["weight_tons"] or r["hh"] or r["ton"] or r["lb"]:
            groups[(r["year_range"], r["from_location"])].append(r)

    docs_dir.mkdir(parents=True, exist_ok=True)
    written = 0

    for (year_range, location), rows in sorted(groups.items()):
        safe_loc = re.sub(r"[^\w\-]", "_", location)
        fname = f"{year_range}_{safe_loc}.md"
        fpath = docs_dir / fname

        # Sort rows by commodity name
        rows.sort(key=lambda r: r["commodity_name"])

        lines = [
            f"# Import Record: {location}, {year_range}",
            "",
            f"**Year range:** {year_range}  ",
            f"**Origin:** {location}  ",
            f"**Destination:** {rows[0]['to_location'] if rows else 'London'}  ",
            "",
            "## Commodities Exported",
            "",
        ]

        for r in rows:
            parts = []
            if r["hh"]:
                parts.append(f"{int(r['hh'])} hogsheads")
            if r["ton"]:
                parts.append(f"{r['ton']:.2f} tons")
            if r["quarter"]:
                parts.append(f"{int(r['quarter'])} quarters")
            if r["lb"]:
                parts.append(f"{int(r['lb'])} lb")
            qty_str = ", ".join(parts) if parts else "quantity unspecified"

            total_str = ""
            if r["total_pounds"]:
                total_str = f" ({int(r['total_pounds'])} total pounds)"
            weight_str = ""
            if r["weight_tons"]:
                weight_str = f" [{r['weight_tons']:.3f} weight tons]"

            lines.append(
                f"- **{r['commodity_name']}**: {qty_str}{total_str}{weight_str}"
            )

        lines += [
            "",
            "## Narrative Summary",
            "",
        ]

        # Build a readable sentence summary for each non-trivial commodity
        commodity_sentences = []
        for r in rows:
            if r["hh"]:
                commodity_sentences.append(
                    f"In {r['year']}, {location} exported {int(r['hh'])} hogsheads of "
                    f"{r['commodity_name']} to {r['to_location']}."
                )
            elif r["weight_tons"] and r["weight_tons"] > 0.01:
                commodity_sentences.append(
                    f"In {r['year']}, {location} exported {r['weight_tons']:.2f} weight tons of "
                    f"{r['commodity_name']} to {r['to_location']}."
                )
            elif r["total_pounds"] and r["total_pounds"] > 0:
                commodity_sentences.append(
                    f"In {r['year']}, {location} exported {int(r['total_pounds'])} pounds of "
                    f"{r['commodity_name']} to {r['to_location']}."
                )

        if commodity_sentences:
            lines.extend(commodity_sentences)
        else:
            lines.append(f"No significant commodity exports recorded for {location} in {year_range}.")

        fpath.write_text("\n".join(lines) + "\n", encoding="utf-8")
        written += 1

    return written


def main():
    parser = argparse.ArgumentParser(description="Parse British Caribbean import record xlsx files.")
    parser.add_argument(
        "--input", default="import_records",
        help="Directory containing .xlsx files (default: import_records)"
    )
    parser.add_argument(
        "--output", default="output/import_records_parsed",
        help="Output directory (default: output/import_records_parsed)"
    )
    args = parser.parse_args()

    input_dir = Path(args.input)
    output_dir = Path(args.output)
    docs_dir = output_dir / "import_docs"

    if not input_dir.exists():
        print(f"ERROR: Input directory '{input_dir}' not found.", file=sys.stderr)
        sys.exit(1)

    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in '{input_dir}'.", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(xlsx_files)} xlsx files in '{input_dir}'")
    print(f"Output directory: '{output_dir}'")
    output_dir.mkdir(parents=True, exist_ok=True)

    # ── Parse all workbooks ────────────────────────────────────────────────────
    all_records: list[dict] = []
    for xlsx_path in tqdm(xlsx_files, desc="Parsing xlsx files"):
        try:
            records = parse_workbook(xlsx_path)
            all_records.extend(records)
        except Exception as e:
            print(f"\nWARN: Failed to parse '{xlsx_path.name}': {e}", file=sys.stderr)

    print(f"\nTotal records parsed: {len(all_records)}")

    # ── Filter out zero-value records for CSV/JSON (keep all for completeness) ─
    nonzero = [r for r in all_records if any([
        r["total_pounds"], r["weight_tons"], r["hh"], r["ton"], r["lb"]
    ])]
    print(f"Non-zero records: {len(nonzero)}")

    # ── Write CSV ──────────────────────────────────────────────────────────────
    csv_path = output_dir / "all_records.csv"
    fieldnames = [
        "file", "year_range", "year", "year_start", "year_end",
        "commodity_name", "from_location", "to_location",
        "ton", "hh", "quarter", "lb", "total_pounds", "weight_tons",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)
    print(f"CSV written: {csv_path} ({len(all_records)} rows)")

    # ── Write JSON ─────────────────────────────────────────────────────────────
    json_path = output_dir / "all_records.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, indent=2, ensure_ascii=False)
    print(f"JSON written: {json_path}")

    # ── Write markdown docs ────────────────────────────────────────────────────
    n_docs = build_markdown_docs(nonzero, docs_dir)
    print(f"Markdown docs written: {n_docs} files in '{docs_dir}'")

    # ── Summary stats ──────────────────────────────────────────────────────────
    commodities = sorted({r["commodity_name"] for r in all_records})
    locations = sorted({r["from_location"] for r in all_records})
    year_starts = sorted({r["year_start"] for r in all_records if r["year_start"] != "unknown"})
    year_range_str = f"{min(year_starts)}–{max(year_starts)}" if year_starts else "unknown"

    print("\n── Summary ──────────────────────────────────────────")
    print(f"  Year range:         {year_range_str}")
    print(f"  Unique commodities: {len(commodities)}")
    print(f"  Unique locations:   {len(locations)}")
    print(f"  Commodities: {', '.join(commodities[:10])}{'...' if len(commodities) > 10 else ''}")
    print(f"  Locations:   {', '.join(locations[:10])}{'...' if len(locations) > 10 else ''}")
    print("─────────────────────────────────────────────────────")


if __name__ == "__main__":
    main()
